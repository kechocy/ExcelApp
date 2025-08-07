import tkinter as tk 
from tkinter import filedialog, messagebox, ttk 
import pandas as pd 
import os 
import sys
import threading
from datetime import datetime
from openpyxl import load_workbook 
from openpyxl.utils.dataframe import dataframe_to_rows
from functools import partial
import zipfile
import xml.etree.ElementTree as ET

class TextRedirector: 
    def __init__(self, widget, tag): 
        self.widget = widget 
        self.tag = tag

    def write(self, message):
        self.widget.config(state='normal')
        self.widget.insert(tk.END, message, (self.tag,))
        self.widget.see(tk.END)
        self.widget.config(state='disabled')

    def flush(self):
        pass

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)
        self.widget.bind("<Motion>", self.motion)

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hidetip()

    def motion(self, event=None):
        self.x = event.x_root + 20
        self.y = event.y_root + 10
        self.schedule()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(500, self.showtip)  # 延迟500ms显示

    def unschedule(self):
        id_ = self.id
        self.id = None
        if id_:
            self.widget.after_cancel(id_)

    def showtip(self):
        if self.tipwindow or not self.text:
            return
        mytext = self.text() if callable(self.text) else self.text
        mytext = str(mytext)
        if not mytext:
            return
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)  # 去掉窗口装饰
        tw.wm_geometry(f"+{self.x}+{self.y}")

        label = tk.Label(tw, text=mytext, justify=tk.LEFT, background="#ffffe0", relief=tk.SOLID, borderwidth=1, font=self.widget.cget("font"))
        label.pack(ipadx=5, ipady=3)

        # 强制重绘一次，获取窗口大小
        tw.update_idletasks()
        tip_width = tw.winfo_width()
        tip_height = tw.winfo_height()

        # 获取主窗口坐标和大小
        root = self.widget.winfo_toplevel()
        root_x = root.winfo_rootx()
        root_y = root.winfo_rooty()
        root_width = root.winfo_width()
        root_height = root.winfo_height()

        # 默认 tooltip 显示位置
        x = self.x
        y = self.y

        # 限制不能超过主窗口右边
        if x + tip_width > root_x + root_width:
            x = root_x + root_width - tip_width - 5  # 留点空隙

        # 限制不能超过主窗口底部
        if y + tip_height > root_y + root_height:
            y = root_y + root_height - tip_height - 5

        # 限制不能太靠左或太靠上
        x = max(x, root_x + 5)
        y = max(y, root_y + 5)

        # 设置 Tooltip 的实际位置
        tw.wm_geometry(f"+{x}+{y}")

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class ExcelMergerApp: 
    def __init__(self, root): 
        self.root = root 
        self.root.title("Excel 合并工具") 
        self.root.geometry("900x600")
        self.root.iconbitmap(self.resource_path("icon.ico"))
        self.template_file = ""
        self.merge_files = []
        

        # 全部使用 grid 布局
        self.root.grid_rowconfigure(2, weight=1) # 第 2 行占比比重
        self.root.grid_rowconfigure(6, weight=2) # 第 6 行占比比重
        self.root.grid_columnconfigure(0, weight=1)

        # 模版文件选择区域
        frame_template = tk.Frame(root)
        frame_template.grid(row=0, column=0, sticky="ew", padx=10, pady=5)
        frame_template.grid_columnconfigure(1, weight=1)

        self.btn_template = tk.Button(frame_template, text="选择模版文件", width=15, anchor='center', command=self.select_template)
        self.btn_template.grid(row=0, column=0, sticky='w')

        self.lbl_template = tk.Label(frame_template, text="", anchor='w', fg='blue')
        self.lbl_template.grid(row=0, column=1, sticky='ew', padx=10)

        self.btn_reset = tk.Button(frame_template, text="重置", width=10, anchor='center', command=self.reset)
        self.btn_reset.grid(row=0, column=2, sticky='e', padx=(0, 20))

        # 合并文件选择区域
        frame_merge_files = tk.Frame(root)
        frame_merge_files.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        frame_merge_files.grid_columnconfigure(1, weight=1)

        self.btn_files = tk.Button(frame_merge_files, text="选择待合并文件", width=15, anchor='center', command=self.select_merge_files)
        self.btn_files.grid(row=0, column=0, sticky='w')

        self.btn_clear_files = tk.Button(frame_merge_files, text="清空", width=10, anchor='center', command=self.clear_merge_files)
        self.btn_clear_files.grid(row=0, column=2, sticky='e', padx=(0, 20))

        # 合并文件列表区域
        frame_file_list = tk.Frame(root)
        frame_file_list.grid(row=2, column=0, sticky="nsew", padx=10)
        frame_file_list.grid_rowconfigure(0, weight=1)
        frame_file_list.grid_columnconfigure(0, weight=1)

        self.file_list = tk.Text(frame_file_list, wrap='none', state='disabled')
        self.file_list.grid(row=0, column=0, sticky='nsew')

        scrollbar_files_y = tk.Scrollbar(frame_file_list, orient=tk.VERTICAL, command=self.file_list.yview)
        scrollbar_files_y.grid(row=0, column=1, sticky='ns')

        scrollbar_files_x = tk.Scrollbar(frame_file_list, orient=tk.HORIZONTAL, command=self.file_list.xview)
        scrollbar_files_x.grid(row=1, column=0, sticky='ew')

        self.file_list.config(yscrollcommand=scrollbar_files_y.set, xscrollcommand=scrollbar_files_x.set)


        # 合并按钮 + 输入信息
        frame_input = tk.Frame(root)
        frame_input.grid(row=4, column=0, sticky="ew", padx=10, pady=5)
        frame_input.columnconfigure(0, weight=1)

        ## 合并按钮
        self.btn_merge = tk.Button(frame_input, text="开始合并", width=15, anchor='center', command=self.start_merge)
        self.btn_merge.grid(row=0, column=0, sticky='w')

        ## 输入框值范围
        ROW_MIN_VAL = 1
        ROW_MAX_VAL = 100
        SHEET_MIN_VAL = 0
        SHEET_MAX_VAL = 1000

        ## 输入框默认值
        self.skip_row_default = "2"
        self.skip_row_var = tk.StringVar(value=self.skip_row_default)
        self.skip_sheet_default = "0"
        self.skip_sheet_var = tk.StringVar(value=self.skip_sheet_default)

        ## 输入框校验
        row_vcmd = (root.register(partial(self.validate_input, min_val=ROW_MIN_VAL, max_val=ROW_MAX_VAL)),"%P")
        sheet_vcmd = (root.register(partial(self.validate_input, min_val=SHEET_MIN_VAL, max_val=SHEET_MAX_VAL)),"%P")

        ## 输入框布局
        skip_sheet_label = tk.Label(frame_input, text="跳过工作表数：")
        self.skip_sheet_spinbox = tk.Spinbox(frame_input, from_=SHEET_MIN_VAL, to=SHEET_MAX_VAL, increment=1, width=10, textvariable=self.skip_sheet_var, validate="key", validatecommand=sheet_vcmd)
        skip_sheet_label.grid(row=0, column=1, sticky='e')
        self.skip_sheet_spinbox.grid(row=0, column=2, sticky='e', padx=(5, 0))
        skip_row_label = tk.Label(frame_input, text="数据起始行：")
        self.skip_row_spinbox = tk.Spinbox(frame_input, from_=ROW_MIN_VAL, to=ROW_MAX_VAL, increment=1, width=10, textvariable=self.skip_row_var, validate="key", validatecommand=row_vcmd)
        skip_row_label.grid(row=0, column=3, sticky='e', padx=(30, 0))
        self.skip_row_spinbox.grid(row=0, column=4, sticky='e', padx=(5, 20))

        ToolTip(skip_sheet_label, "合并时跳过模板文件中前几个 Sheet ？")
        ToolTip(skip_row_label, "Sheet 中从第几行开始是需要合并的有效数据？")

        self.skip_row_spinbox.bind("<FocusOut>", lambda e: self.out_validate_input(self.skip_row_var, self.skip_row_default))
        self.skip_sheet_spinbox.bind("<FocusOut>", lambda e: self.out_validate_input(self.skip_sheet_var, self.skip_sheet_default))

        # 输出框标题
        label_output = tk.Label(root, text="输出信息：")
        label_output.grid(row=5, column=0, sticky='w', padx=10)

        # 输出信息区域
        frame_output = tk.Frame(root)
        frame_output.grid(row=6, column=0, sticky="nsew", padx=10, pady=5)
        frame_output.grid_rowconfigure(0, weight=1)
        frame_output.grid_columnconfigure(0, weight=1)

        self.output_box = tk.Text(frame_output, wrap='none', bg="#f0f0f0", state='disabled')
        self.output_box.grid(row=0, column=0, sticky="nsew")

        scrollbar_y = tk.Scrollbar(frame_output, orient=tk.VERTICAL, command=self.output_box.yview)
        scrollbar_y.grid(row=0, column=1, sticky="ns")

        scrollbar_x = tk.Scrollbar(frame_output, orient=tk.HORIZONTAL, command=self.output_box.xview)
        scrollbar_x.grid(row=1, column=0, sticky="ew")

        self.output_box.config(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        # 进度条
        self.frame_progress = tk.Frame(root)
        
        self.frame_progress.grid_columnconfigure(0, weight=3)
        self.progress = ttk.Progressbar(self.frame_progress, orient='horizontal', mode='determinate')
        self.progress_label = tk.Label(self.frame_progress, text="")
        self.progress.grid(row=0, column=0, sticky='ew')
        self.progress_label.grid(row=0, column=1, sticky='e', padx=(5, 20))

        # 底部版权标签
        copyright_label = tk.Label(root, text="© 2025 Kecho", fg="gray", font=("Arial", 8))
        copyright_label.grid(row=8, column=0, sticky="w", pady=2, padx=10)

        # 重定向 stdout
        sys.stdout = TextRedirector(self.output_box, "stdout")
        # sys.stderr = TextRedirector(self.output_box, "stderr")

        # 按钮状态初始化
        self.init_set_buttons()

    def select_template(self):
        file = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        if file:
            self.template_file = file
            basename = os.path.basename(file)
            self.lbl_template.config(text=basename)
            self.lbl_template.fullpath = file
            if hasattr(self.lbl_template, "fullpath") and not hasattr(self.lbl_template, "_has_tooltip"):
                ToolTip(self.lbl_template, text=lambda: self.lbl_template.fullpath)
                self.lbl_template._has_tooltip = True
            if self.template_file:
                self.skip_row_spinbox.config(state='normal')
                self.skip_sheet_spinbox.config(state='normal')
                if len(self.merge_files) >= 2:
                    self.btn_merge.config(state='normal')

    def select_merge_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel 文件", "*.xls *.xlsx")])
        
        if files:
            for f in files:
                if f not in self.merge_files:
                    self.merge_files.append(f)
            self.update_file_list()
            if self.merge_files:
                self.btn_clear_files.config(state='normal')
            else:
                self.btn_clear_files.config(state='disabled')
            if len(self.merge_files) >= 2 and self.template_file:
                self.btn_merge.config(state='normal')
            else:
                self.btn_merge.config(state='disabled')

    def clear_merge_files(self):
        self.merge_files = []
        self.update_file_list()
        self.btn_clear_files.config(state='disabled')
        self.btn_merge.config(state='disabled')
    
    def reset(self):
        self.template_file = ""
        self.merge_files = []
        self.update_file_list()
        self.lbl_template.config(text="")
        if hasattr(self.lbl_template, "fullpath"):
            self.lbl_template.fullpath = ""
        if self.skip_row_spinbox['state'] == "normal" and not self.skip_row_var.get() == self.skip_row_default:
            self.skip_row_var.set(self.skip_row_default)
        if self.skip_sheet_spinbox['state'] == "normal" and not self.skip_sheet_var.get == self.skip_sheet_default:
            self.skip_sheet_var.set(self.skip_sheet_default)
        self.init_set_buttons()

    def update_file_list(self):
        self.file_list.config(state='normal')
        self.file_list.delete(1.0, tk.END) # 清空内容
        # if self.merge_files:
        #     self.file_list.insert(tk.END, "待合并文件:\n")
        for f in self.merge_files:
            self.file_list.insert(tk.END, f"  {f}\n")
        self.file_list.config(state='disabled')

    def start_merge(self):

        self.start_progress()
        self.set_buttons_state('disabled')

        thread = threading.Thread(target=self.run_merge_thread, daemon=True)
        thread.start()

    def run_merge_thread(self):
        try:
            result = self.merge_into_template()

            if "skip_sheet_error" == result:
                messagebox.showerror("错误", f"请正确设置 “跳过工作表数” ")
            elif "template_type_error" == result:
                messagebox.showerror("错误", f"请选择支持的文件格式的模板文件")
            elif "file_type_error" == result:
                messagebox.showerror("错误", f"请选择支持的文件格式的待合并文件")
            elif result is not None:
                messagebox.showinfo("完成", f"合并完成。已保存为：\n{result}")        
            else:
                messagebox.showerror("错误", f"请选择正确的模板文件")
        except Exception as e:
           messagebox.showerror("错误", f"合并失败: {e}")
            
        finally:
            self.root.after(0, lambda: self.remove_progress())
            self.set_buttons_state('normal')

    def set_buttons_state(self, state):
        self.btn_reset.config(state=state)
        self.btn_template.config(state=state)
        self.btn_clear_files.config(state=state)
        self.btn_files.config(state=state)
        self.btn_merge.config(state=state)
        self.skip_row_spinbox.config(state=state)
        self.skip_sheet_spinbox.config(state=state)
    
    def init_set_buttons(self):
        self.btn_template.config(state='normal')
        self.btn_clear_files.config(state='disabled')
        self.btn_files.config(state='normal')
        self.btn_merge.config(state='disabled')
        self.skip_row_spinbox.config(state='disabled')
        self.skip_sheet_spinbox.config(state='disabled')

    def start_progress(self):
        self.progress_label.config(text="0%")
        self.progress["value"] = 0
        if not self.frame_progress.winfo_ismapped():
            self.frame_progress.grid(row=7, column=0, sticky="ew", padx=10, pady=5)

    def remove_progress(self):
        self.frame_progress.grid_forget()

    def validate_input(self, input_value, min_val, max_val): # 校验 spinbox 输入合法
        if input_value == "":
            return True # 允许删除内容，暂时为空
        if input_value.isdigit(): # 非负整数
           if input_value !="0" and input_value.startswith("0"):
               return False # 不允许前置 0，除非是 0
           
           val = int(input_value)
           return min_val <= val <= max_val
        return False
    
    def out_validate_input(self, var, default): # 为空则设置默认值
        val = var.get()
        if val == "":
            var.set(default)

    def resource_path(self, relative_path):
        if hasattr(sys, "_MEIPASS"):
            return os.path.join(sys._MEIPASS, relative_path)
        return relative_path
    
    def merge_into_template(self):
        # 获取输入值
        skip_sheet = int(self.skip_sheet_var.get())
        skip_row = int(self.skip_row_var.get())
    
        print("正在校验文件格式...")
        if self.is_strict_openxml(self.template_file):
            print(f"不支持 Strict Open XML 格式的 Excel 文件，请将模板文件另存为标准 .xlsx 格式")
            return "template_type_error"
        elif self.template_file.endswith('.xls'):
            print(f"模板文件不支持 xls 格式，请另存为标准 .xlsx 格式")
            return "template_type_error"
        for file in self.merge_files:
            if self.is_strict_openxml(file):
                print(f"不支持 Strict Open XML 格式的 Excel 文件，请将 {file} 另存为标准 .xlsx 格式")
                return "file_type_error"
        
        print("校验成功，正在读取模板文件...")
        wb = load_workbook(self.template_file)
        
        sheet_names = wb.sheetnames
        sheet_names_set = set(sheet_names)

        sheet_count = len(sheet_names)
        if sheet_count < 1:
            print(f"模板文件中至少应该有 1 个 Sheet")
            return
        
        if skip_sheet >= len(sheet_names):
            print(f"模板文件中只有 {sheet_count} 个 Sheet，无法跳过 {skip_sheet} 个 Sheet")
            return "skip_sheet_error"
        
        need_deal_sheet_names = sheet_names[skip_sheet:]  # 跳过指定数量个 Sheet
        need_deal_sheet_count = len(need_deal_sheet_names)
        self.progress["maximum"] = need_deal_sheet_count
        print(f"共有 {need_deal_sheet_count} 个 Sheet 等待处理")

        print("正在校验子文件中所有 Sheet 是否在模板中...")
        for file in self.merge_files:
            engine = "xlrd" if file.endswith(".xls") else "openpyxl"
            try:
                sub_sheet_names = pd.ExcelFile(file, engine=engine).sheet_names
                for s in sub_sheet_names:
                    if s not in sheet_names_set:
                        print(f"模板中没有 {os.path.basename(file)} 中的 Sheet: {s}，请选择正确的模板")
                        return
            except Exception as e:
                print(f"读取失败 {os.path.basename(file)} 失败。原因: {e}")

        print("校验通过，开始处理...")

        # 开始合并
        for idx, sheet in enumerate(need_deal_sheet_names):
            print(f"[{idx + 1}/{need_deal_sheet_count}] 处理: {sheet}")
            combined = []

            for file in self.merge_files:
                engine = "xlrd" if file.endswith(".xls") else "openpyxl"
                try:
                    excel = pd.ExcelFile(file, engine=engine)

                    if sheet in excel.sheet_names:
                        df = pd.read_excel(file, sheet_name=sheet, skiprows=(skip_row - 1), header=None, engine=engine)
                        df = df.dropna(how='all') # 去除空行
                        # if not df.empty:
                        #     df = df[df.iloc[:,0].notna()] # 只取第一列非空
                        if not df.empty:
                            combined.append(df)

                except Exception as e:
                    print(f"读取失败: {os.path.basename(file)} 的 {sheet}，跳过。原因: {e}")

            if combined:
                df_all = pd.concat(combined, ignore_index=True)
                ws = wb[sheet]
                if ws.max_row > (skip_row - 1):
                    ws.delete_rows(skip_row, ws.max_row - skip_row + 1)
                for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=False), start=skip_row):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

            self.progress["value"] = idx + 1
            percent = ((idx + 1) / (need_deal_sheet_count + 1)) * 100 # 分母比实际大一点，因为最后保存文件也需要时间
            self.progress_label.config(text=f"{percent:.1f}%")
            self.root.update_idletasks()
        
        out_file = f"{os.path.splitext(os.path.basename(self.template_file))[0]}_合并_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(out_file)
        print(f"合并完成，已保存为 {out_file}\n")
        self.progress_label.config(text="100%") # 手动调整显示文字
        return out_file
    
    def is_strict_openxml(self, file_path):
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                content = z.read('xl/workbook.xml')
                root = ET.fromstring(content)
                ns = root.tag.split('}')[0].strip('{')
                if ns == 'http://purl.oclc.org/ooxml/spreadsheetml/main':
                    return True
                else:
                    return False
        except Exception:
            return False
        
# class ExcelReader:
#     def __init__(self, file_path):
#         self.file_path = file_path
#         self.ext = os.path.splitext(file_path)[1].lower() # 文件的后缀名
#         if self.ext == '.xls':
#             self.engine = 'xlrd'
#             self.wb = xlrd.open_workbook(file_path)
#         elif self.ext == '.xlsx':
#             self.engine = 'openpyxl'
#             self.wb = load_workbook(file_path)
#         else:
#             raise ValueError("只支持 .xls 和 .xlsx 格式的文件")

#     def sheet_names(self):
#         if self.engine == 'xlrd':
#             return self.wb.sheet_names()
#         else:
#             return self.wb.sheetnames
    
#     # 通过 sheet 索引(int)或名称(str)获取 sheet 对象
#     def sheet(self, identifier):
#         if self.engine == 'xlrd':
#             if isinstance(identifier, int):
#                 return self.wb.sheet_by_index(identifier)
#             elif isinstance(identifier, str):
#                 return self.wb.sheet_by_name(identifier)
#             else:
#                 raise ValueError("获取 Sheet 名称时参数必须是 int 或 str")
#         else:
#             if isinstance(identifier, int):
#                 name = self.wb.sheetnames[identifier]
#                 return self.wb[name]
#             elif isinstance(identifier, str):
#                 return self.wb[identifier]
#             else:
#                 raise ValueError("获取 Sheet 名称时参数必须是 int 或 str")
            
#     def nrows(self, sheet):
#         if self.engine == 'xlrd':
#             return sheet.nrows
#         else:
#             return sheet.max_row
        
#     def ncols(self, sheet):
#         if self.engine == 'xlrd':
#             return sheet.ncols
#         else:
#             return sheet.max_column

#     # 读取单元格的值，xlrd 行列从 0 开始，openpyxl 行列从 1 开始
#     def cell_value(self, sheet, row, col):
#         if self.engine == 'xlrd':
#             return sheet.cell_value(row, col)
#         else:
#             return sheet.cell(row=row + 1, column=col + 1).value
# 启动 GUI
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMergerApp(root)
    root.mainloop()